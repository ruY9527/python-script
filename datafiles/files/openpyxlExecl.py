import openpyxl


def writeDataByOpenpyxl():
    data = openpyxl.Workbook() # 新建工作簿
    data.create_sheet('Sheet1') # 添加页
    table =  data.active
    table.cell(1,1,'TestBaoYang')
    data.save('/home/luohong/Downloads/222.xlsx') 
 
def writeDataByOpenpyxl():
    data = openpyxl.load_workbook('/home/luohong/Downloads/111.xlsx') # 读取xlsx文件
    table = data.get_sheet_by_name('Over') # 获得指定名称的页
    nrows = table.rows
    ncols = table.columns
    print(type(nrows))
    for row in nrows:
        print(row)
        line = [col.value for col in row]
        print(line)
    print(table.cell(1,1).value)

if __name__ == '__main__':
    writeDataByOpenpyxl()